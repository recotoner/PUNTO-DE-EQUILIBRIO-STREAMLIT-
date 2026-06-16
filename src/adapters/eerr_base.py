"""Shared parsing behavior for EERR spreadsheets."""

from __future__ import annotations

import calendar
import re
from datetime import date

import pandas as pd
from openpyxl import load_workbook

from src.adapters.common import ExcelSource, clave_texto, normalizar_texto
from src.models import EERRNormalizado, Periodo

MONTHS = {
    "ENE": 1,
    "ENERO": 1,
    "FEB": 2,
    "FEBRERO": 2,
    "MAR": 3,
    "MARZO": 3,
    "ABR": 4,
    "ABRIL": 4,
    "MAY": 5,
    "MAYO": 5,
    "JUN": 6,
    "JUNIO": 6,
    "JUL": 7,
    "JULIO": 7,
    "AGO": 8,
    "AGOSTO": 8,
    "SEP": 9,
    "SEPT": 9,
    "SEPTIEMBRE": 9,
    "OCT": 10,
    "OCTUBRE": 10,
    "NOV": 11,
    "NOVIEMBRE": 11,
    "DIC": 12,
    "DICIEMBRE": 12,
}


class EERRBaseAdapter:
    formato = ""

    def load(self, source: ExcelSource) -> EERRNormalizado:
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            empresa = normalizar_texto(sheet["A2"].value)
            rut = normalizar_texto(sheet["A3"].value)
            period_text = normalizar_texto(sheet["A4"].value)
            periodo = self._parse_period(period_text)
            month_columns = self._month_columns(sheet)
            total_column = self._total_column(sheet)
            rows = self._parse_rows(sheet, month_columns, total_column)
        finally:
            workbook.close()

        cuentas = pd.DataFrame(rows)
        totales = {
            row["concepto_clave"]: float(row["total"])
            for row in rows
            if row["concepto_clave"].startswith("TOTAL ")
            or row["concepto_clave"] in {"MARGEN DE EXPLOTACION", "RESULTADO OPERACIONAL"}
        }
        return EERRNormalizado(
            empresa=empresa,
            rut=rut,
            periodo=periodo,
            formato=self.formato,
            cuentas=cuentas,
            totales=totales,
            meses=[name for _, name in month_columns],
            metadata={"periodo_texto": period_text},
        )

    @staticmethod
    def _parse_period(text: str) -> Periodo:
        normalized = clave_texto(text)
        year_match = re.search(r"\b(20\d{2})\b", normalized)
        if not year_match:
            raise ValueError("No fue posible detectar el año del EERR.")
        year = int(year_match.group(1))
        found = [number for name, number in MONTHS.items() if re.search(rf"\b{name}\b", normalized)]
        if not found:
            raise ValueError("No fue posible detectar los meses del EERR.")
        start_month, end_month = min(found), max(found)
        last_day = calendar.monthrange(year, end_month)[1]
        return Periodo(date(year, start_month, 1), date(year, end_month, last_day))

    @staticmethod
    def _month_columns(sheet) -> list[tuple[int, str]]:
        columns: list[tuple[int, str]] = []
        for column in range(3, sheet.max_column + 1):
            label = clave_texto(sheet.cell(5, column).value)
            if label in MONTHS:
                columns.append((column, normalizar_texto(sheet.cell(5, column).value)))
        return columns

    @staticmethod
    def _total_column(sheet) -> int:
        for column in range(3, sheet.max_column + 1):
            if clave_texto(sheet.cell(5, column).value) == "TOTAL":
                return column
        raise ValueError("El EERR no contiene una columna TOTAL.")

    @staticmethod
    def _parse_rows(sheet, month_columns: list[tuple[int, str]], total_column: int) -> list[dict]:
        rows = []
        for row_number in range(6, sheet.max_row + 1):
            section = normalizar_texto(sheet.cell(row_number, 1).value)
            account = normalizar_texto(sheet.cell(row_number, 2).value)
            concept = account or section
            total = sheet.cell(row_number, total_column).value
            if not concept or total is None:
                continue
            monthly = {
                name: float(sheet.cell(row_number, column).value or 0)
                for column, name in month_columns
            }
            rows.append(
                {
                    "fila_origen": row_number,
                    "seccion": section,
                    "cuenta": account,
                    "concepto": concept,
                    "concepto_clave": clave_texto(concept),
                    "total": float(total),
                    "valores_mensuales": monthly,
                }
            )
        return rows
