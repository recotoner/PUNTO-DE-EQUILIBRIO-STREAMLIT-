"""Facade that detects and loads the supported EERR formats."""

from __future__ import annotations

from openpyxl import load_workbook

from src.adapters.common import ExcelSource, clave_texto
from src.adapters.eerr_acumulado_adapter import EERRAcumuladoAdapter
from src.adapters.eerr_mensual_adapter import EERRMensualAdapter
from src.models import EERRNormalizado


def detectar_formato_eerr(source: ExcelSource) -> str:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = [
            clave_texto(sheet.cell(5, column).value)
            for column in range(3, sheet.max_column + 1)
        ]
    finally:
        workbook.close()

    non_total_headers = [header for header in headers if header and header != "TOTAL"]
    return "mensual" if non_total_headers else "acumulado"


def cargar_eerr(source: ExcelSource) -> EERRNormalizado:
    formato = detectar_formato_eerr(source)
    if hasattr(source, "seek"):
        source.seek(0)
    if formato == "mensual":
        return EERRMensualAdapter().load(source)
    return EERRAcumuladoAdapter().load(source)
